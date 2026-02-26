from rest_framework import viewsets, status
from rest_framework.response import Response
from rest_framework.decorators import action
from django.utils.dateparse import parse_date
from django.http import HttpResponse
from io import BytesIO
from django.conf import settings
from openpyxl import Workbook, load_workbook
from decimal import Decimal
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from urllib.parse import quote

from .models import (
    Role, Permission, User,
    PassengerCar, NormsPassengerCars, PassengerCarWaybill,
    PassengerCarWaybillRecord, OdometerFuelPassengerCar,
    FireTruck, NormsFireTruck, FireTruckWaybill,
    FireTruckWaybillRecord, OdometerFuelFireTruck,
)
from .serializers import (
    RoleSerializer, PermissionSerializer, UserSerializer,
    PassengerCarSerializer, NormsPassengerCarsSerializer,
    PassengerCarWaybillSerializer, PassengerCarWaybillRecordSerializer,
    OdometerFuelPassengerCarSerializer,
    FireTruckSerializer, NormsFireTruckSerializer,
    FireTruckWaybillSerializer, FireTruckWaybillRecordSerializer,
    OdometerFuelFireTruckSerializer,
)

class SoftDeleteModelViewSet(viewsets.ModelViewSet):
    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        instance.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)
        instance = self.get_object()
        instance.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)

# --- Роли, права и пользователи ---
class RoleViewSet(SoftDeleteModelViewSet):
    queryset = Role.objects.all()
    serializer_class = RoleSerializer

class PermissionViewSet(SoftDeleteModelViewSet):
    queryset = Permission.objects.all()
    serializer_class = PermissionSerializer

class UserViewSet(SoftDeleteModelViewSet):
    queryset = User.objects.all()
    serializer_class = UserSerializer

# --- Легковые ---

class PassengerCarViewSet(SoftDeleteModelViewSet):
    queryset = PassengerCar.objects.all()
    serializer_class = PassengerCarSerializer

class NormsPassengerCarsViewSet(SoftDeleteModelViewSet):
    queryset = NormsPassengerCars.objects.all()
    serializer_class = NormsPassengerCarsSerializer

    @action(detail=False, methods=['get'], url_path='for-date')
    def for_date(self, request):
        """
        GET /api/passenger-car-norms/for-date/?car=<id>&season=<summer|winter>&date=YYYY-MM-DD

        Возвращает последнюю норму для указанной машины и сезона
        на дату документа (date__lte=дата, сортировка по дате/ID).
        """
        car_id = request.query_params.get('car')
        season = request.query_params.get('season')
        date_str = request.query_params.get('date')

        if not car_id or not season or not date_str:
            return Response(
                {"detail": "Параметры car, season и date обязательны"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        doc_date = parse_date(date_str)
        if not doc_date:
            return Response(
                {"detail": "Неверный формат date, ожидается YYYY-MM-DD"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        norm = (
            NormsPassengerCars.objects
            .filter(car_id=car_id, season=season, date__lte=doc_date)
            .order_by('-date', '-id')
            .first()
        )
        if not norm:
            return Response(
                {"detail": "Норма не найдена"},
                status=status.HTTP_404_NOT_FOUND,
            )

        serializer = self.get_serializer(norm)
        return Response(serializer.data)

class OdometerFuelPassengerCarViewSet(SoftDeleteModelViewSet):
    queryset = OdometerFuelPassengerCar.objects.all()
    serializer_class = OdometerFuelPassengerCarSerializer

    @action(detail=False, methods=['get'], url_path='last')
    def last_record(self, request):
        """
        GET /api/passenger-car-odometer-fuel/last/?car=<id>

        Возвращает последнюю запись по данному автомобилю
        (по дате и ID).
        """
        car_id = request.query_params.get('car')
        if not car_id:
            return Response(
                {"detail": "Параметр car обязателен"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        obj = (
            OdometerFuelPassengerCar.objects
            .filter(car_id=car_id)
            .order_by('-date', '-id')
            .first()
        )
        if not obj:
            return Response(
                {"detail": "Записей не найдено"},
                status=status.HTTP_404_NOT_FOUND,
            )

        serializer = self.get_serializer(obj)
        return Response(serializer.data)

class PassengerCarWaybillViewSet(SoftDeleteModelViewSet):
    queryset = PassengerCarWaybill.objects.all()
    serializer_class = PassengerCarWaybillSerializer

    @action(detail=False, methods=['get'], url_path='export-excel')
    def export_excel(self, request):
        """
        GET /api/passenger-car-waybills/export-excel/?car=<id>&from=YYYY-MM-DD&to=YYYY-MM-DD

        Выгрузка эксплуатационной карточки легкового автомобиля по Excel-шаблону.
        """
        car_id = request.query_params.get('car')
        from_str = request.query_params.get('from')
        to_str = request.query_params.get('to')

        if not car_id or not from_str or not to_str:
            return Response(
                {"detail": "Параметры car, from и to обязательны"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        from_date = parse_date(from_str)
        to_date = parse_date(to_str)
        if not from_date or not to_date:
            return Response(
                {"detail": "Неверный формат дат, используйте YYYY-MM-DD"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        records = (
            PassengerCarWaybillRecord.objects
            .filter(
                passenger_car_waybill__car_id=car_id,
                passenger_car_waybill__date__gte=from_date,
                passenger_car_waybill__date__lte=to_date,
            )
            .select_related('passenger_car_waybill__driver',
                            'passenger_car_waybill__car')
            .order_by('passenger_car_waybill__date', 'id')
        )

        if not records.exists():
            return Response(
                {"detail": "Записей за указанный период не найдено"},
                status=status.HTTP_404_NOT_FOUND,
            )

        car = records.first().passenger_car_waybill.car

        # ----- открываем шаблон -----
        template_path = settings.BASE_DIR / 'report_templates' / 'passenger_car.xlsx'
        wb = load_workbook(template_path)
        ws = wb.active  

        ws['D2'] = car.number
        ws['I2'] = from_date.strftime('%d.%m.%Y')
        ws['N2'] = to_date.strftime('%d.%m.%Y')

        data_start_row = 7
        row_idx = data_start_row

        total_distance_city = 0
        total_distance_area = 0
        total_distance = 0
        total_fuel_used_city = Decimal('0.000')
        total_fuel_used_area = Decimal('0.000')
        total_fuel_used_fact = Decimal('0.000')
        total_fuel_used_normal = Decimal('0.000')
        total_fuel_refueled = Decimal('0.000')
        total_savings = 0.0
        total_overrun = 0.0

        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000'),
        )

        yellow_fill = PatternFill(
            start_color="FFFF00",
            end_color="FFFF00",
            fill_type="solid",
        )

        green_fill = PatternFill(
            start_color="92D050",
            end_color="92D050",
            fill_type="solid",
        )

        red_fill = PatternFill(
            start_color="FF0000",
            end_color="FF0000",
            fill_type="solid",
        )

        for rec in records:
            wb_obj = rec.passenger_car_waybill
            driver = wb_obj.driver

            savings = 0.0
            overrun = 0.0

            fio = f"{driver.surname} {driver.name[0]}. {driver.last_name[0]}."

            ws.cell(row=row_idx, column=1, value=wb_obj.date.strftime('%d.%m.%Y'))
            ws.cell(row=row_idx, column=2, value=fio)
            cell = ws.cell(row=row_idx, column=3, value=rec.fuel_before_departure)
            cell.fill = yellow_fill
            cell = ws.cell(row=row_idx, column=4, value=rec.odometer_before) 
            cell.fill = yellow_fill
            ws.cell(row=row_idx, column=5, value=rec.distance_total_km)
            ws.cell(row=row_idx, column=6, value=rec.distance_city_km)
            ws.cell(row=row_idx, column=7, value=rec.distance_area_km)
            ws.cell(row=row_idx, column=8, value=rec.fuel_used_city)
            ws.cell(row=row_idx, column=9, value=rec.fuel_used_area)
            cell = ws.cell(row=row_idx, column=10, value=rec.fuel_used_normal) 
            cell.fill = yellow_fill
            cell = ws.cell(row=row_idx, column=11, value=rec.fuel_used) 
            cell.fill = yellow_fill
            cell = ws.cell(row=row_idx, column=12, value=rec.fuel_refueled)
            cell.fill = green_fill
            cell = ws.cell(row=row_idx, column=13, value=rec.fuel_on_return)
            cell.fill = yellow_fill
            cell = ws.cell(row=row_idx, column=14, value=rec.odometer_after)
            cell.fill = yellow_fill
            if(rec.fuel_used_normal > rec.fuel_used):
                savings = float(rec.fuel_used_normal - rec.fuel_used)
                overrun = 0.0
                cell = ws.cell(row=row_idx, column=15, value=savings)
                cell.fill = green_fill
                cell = ws.cell(row=row_idx, column=16, value=overrun)
                cell.fill = red_fill
            elif(rec.fuel_used_normal < rec.fuel_used):
                savings = 0.0
                overrun = float(rec.fuel_used - rec.fuel_used_normal)
                cell = ws.cell(row=row_idx, column=16, value=overrun)
                cell.fill = red_fill
                cell = ws.cell(row=row_idx, column=15, value=savings)
                cell.fill = green_fill
            else:
                cell = ws.cell(row=row_idx, column=16, value=0)
                cell.fill = red_fill
                cell = ws.cell(row=row_idx, column=15, value=0)
                cell.fill = green_fill                  
            
            total_distance_city += rec.distance_city_km
            total_distance_area += rec.distance_area_km
            total_distance += rec.distance_total_km
            total_fuel_used_city += rec.fuel_used_city
            total_fuel_used_area += rec.fuel_used_area
            total_fuel_used_fact += rec.fuel_used
            total_fuel_refueled += rec.fuel_refueled
            total_savings += savings
            total_overrun += overrun
            total_fuel_used_normal += rec.fuel_used_normal

            row_idx += 1
        
        cell = ws.cell(row=row_idx, column=2, value="ИТОГО")
        cell.fill = yellow_fill
        cell = ws.cell(row=row_idx, column=5, value=total_distance)
        cell.fill = yellow_fill
        cell = ws.cell(row=row_idx, column=6, value=total_distance_city)
        cell.fill = yellow_fill
        cell = ws.cell(row=row_idx, column=7, value=total_distance_area)
        cell.fill = yellow_fill
        cell = ws.cell(row=row_idx, column=8, value=float(total_fuel_used_city))
        cell.fill = yellow_fill
        cell = ws.cell(row=row_idx, column=9, value=float(total_fuel_used_area))
        cell.fill = yellow_fill
        cell = ws.cell(row=row_idx, column=10, value=float(total_fuel_used_normal))
        cell.fill = yellow_fill
        cell = ws.cell(row=row_idx, column=11, value=float(total_fuel_used_fact))
        cell.fill = yellow_fill
        cell = ws.cell(row=row_idx, column=12, value=float(total_fuel_refueled))
        cell.fill = green_fill
        cell = ws.cell(row=row_idx, column=15, value=float(total_savings))
        cell.fill = green_fill
        cell = ws.cell(row=row_idx, column=16, value=float(total_overrun))
        cell.fill = red_fill

        data_end_row = row_idx
        for r in range(data_start_row, data_end_row + 1):
            for c in range(1, 16 + 1):
                cell = ws.cell(row=r, column=c)
                cell.border = thin_border
                cell.font = Font(name='Times New Roman', size=11)
                cell.alignment = Alignment(horizontal='center', vertical='center')

        for c in range(1, 16 + 1):
            cell = ws.cell(row=data_end_row, column=c)
            cell.font = Font(name='Times New Roman', size=11, bold=True)

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"Путевые листы легкового автомобиля({car.number}) за период {from_date.strftime('%d.%m.%Y')}-{to_date.strftime('%d.%m.%Y')}.xlsx"
        response = HttpResponse(
            output.read(),
            content_type=("application/vnd.openxmlformats-officedocument."
                          "spreadsheetml.sheet"),
        )
        quoted_filename = quote(filename)
        response['Content-Disposition'] = f"attachment; filename*=UTF-8''{quoted_filename}"
        return response

class PassengerCarWaybillRecordViewSet(SoftDeleteModelViewSet):
    queryset = PassengerCarWaybillRecord.objects.select_related('passenger_car_waybill')
    serializer_class = PassengerCarWaybillRecordSerializer

# --- Пожарные ---

class FireTruckViewSet(SoftDeleteModelViewSet):
    queryset = FireTruck.objects.all()
    serializer_class = FireTruckSerializer

class NormsFireTruckViewSet(SoftDeleteModelViewSet):
    queryset = NormsFireTruck.objects.all()
    serializer_class = NormsFireTruckSerializer

    @action(detail=False, methods=['get'], url_path='for-date')
    def for_date(self, request):
        """
        GET /api/fire-truck-norms/for-date/?car=<id>&season=<summer|winter>&date=YYYY-MM-DD

        Возвращает последнюю норму для указанного ПА и сезона
        на дату документа (date__lte=дата).
        """
        car_id = request.query_params.get('car')
        season = request.query_params.get('season')
        date_str = request.query_params.get('date')

        if not car_id or not season or not date_str:
            return Response(
                {"detail": "Параметры car, season и date обязательны"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        doc_date = parse_date(date_str)
        if not doc_date:
            return Response(
                {"detail": "Неверный формат date, ожидается YYYY-MM-DD"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        norm = (
            NormsFireTruck.objects
            .filter(car_id=car_id, season=season, date__lte=doc_date)
            .order_by('-date', '-id')
            .first()
        )
        if not norm:
            return Response(
                {"detail": "Норма не найдена"},
                status=status.HTTP_404_NOT_FOUND,
            )

        serializer = self.get_serializer(norm)
        return Response(serializer.data)

class OdometerFuelFireTruckViewSet(SoftDeleteModelViewSet):
    queryset = OdometerFuelFireTruck.objects.all()
    serializer_class = OdometerFuelFireTruckSerializer

    @action(detail=False, methods=['get'], url_path='last')
    def last_record(self, request):
        """
        GET /api/fire-truck-odometer-fuel/last/?car=<id>

        Возвращает последнюю запись по данному пожарному автомобилю.
        """
        car_id = request.query_params.get('car')
        if not car_id:
            return Response(
                {"detail": "Параметр car обязателен"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        obj = (
            OdometerFuelFireTruck.objects
            .filter(car_id=car_id)
            .order_by('-date', '-id')
            .first()
        )
        if not obj:
            return Response(
                {"detail": "Записей не найдено"},
                status=status.HTTP_404_NOT_FOUND,
            )

        serializer = self.get_serializer(obj)
        return Response(serializer.data)

class FireTruckWaybillViewSet(SoftDeleteModelViewSet):
    queryset = FireTruckWaybill.objects.all()
    serializer_class = FireTruckWaybillSerializer

    @action(detail=False, methods=['get'], url_path='export-excel')
    def export_excel(self, request):
        """
        GET /api/fire-truck-waybills/export-excel/?car=<id>&from=YYYY-MM-DD&to=YYYY-MM-DD

        Выгрузка эксплуатационной карточки пожарного автомобиля по Excel-шаблону
        + строка итогов.
        """
        car_id = request.query_params.get('car')
        from_str = request.query_params.get('from')
        to_str = request.query_params.get('to')

        if not car_id or not from_str or not to_str:
            return Response(
                {"detail": "Параметры car, from и to обязательны"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        from_date = parse_date(from_str)
        to_date = parse_date(to_str)
        if not from_date or not to_date:
            return Response(
                {"detail": "Неверный формат дат, используйте YYYY-MM-DD"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        records = (
            FireTruckWaybillRecord.objects
            .filter(
                fire_truck_waybill__car_id=car_id,
                fire_truck_waybill__date__gte=from_date,
                fire_truck_waybill__date__lte=to_date,
            )
            .select_related('fire_truck_waybill__driver',
                            'fire_truck_waybill__car')
            .order_by('fire_truck_waybill__date', 'id')
        )

        if not records.exists():
            return Response(
                {"detail": "Записей за указанный период не найдено"},
                status=status.HTTP_404_NOT_FOUND,
            )

        car = records.first().fire_truck_waybill.car

        # ----- открываем шаблон -----
        template_path = settings.BASE_DIR / 'report_templates' / 'fire_truck.xlsx'
        wb = load_workbook(template_path)
        ws = wb.active  # или wb['Имя_листа']

        # Шапка (подстрои адреса под свой шаблон)
        ws['D2'] = car.number
        ws['K2'] = from_date.strftime('%d.%m.%Y')
        ws['R2'] = to_date.strftime('%d.%m.%Y')

        data_start_row = 7
        row_idx = data_start_row

        # Итоговые суммы
        total_distance_km = 0
        total_time_with_pump = 0
        total_time_without_pump = 0
        total_fuel_by_distance = Decimal('0.000')
        total_fuel_with_pump = Decimal('0.000')
        total_fuel_without_pump = Decimal('0.000')
        total_fuel_normal = Decimal('0.000')
        total_fuel_fact = Decimal('0.000')
        total_fuel_refueled = Decimal('0.000')
        total_savings = Decimal('0.000')
        total_overrun = Decimal('0.000')

        # Стили
        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000'),
        )

        yellow_fill = PatternFill(
            start_color="FFFF00",
            end_color="FFFF00",
            fill_type="solid",
        )

        green_fill = PatternFill(
            start_color="92D050",
            end_color="92D050",
            fill_type="solid",
        )

        red_fill = PatternFill(
            start_color="FF0000",
            end_color="FF0000",
            fill_type="solid",
        )

        center_font = Font(name='Times New Roman', size=11)
        bold_font = Font(name='Times New Roman', size=11, bold=True)
        center_align = Alignment(horizontal='center', vertical='center')

        for rec in records:
            wb_obj = rec.fire_truck_waybill
            driver = wb_obj.driver

            # Наименование и место работы = target + driving_route
            # Если driving_route нет/пустой — просто target
            route = getattr(rec, 'driving_route', '') or ''
            name_place = (rec.target or '') + (f" {route}" if route else '')

            # Экономия/перерасход
            savings = Decimal('0.000')
            overrun = Decimal('0.000')

            fio = f"{driver.surname} {driver.name[0]}. {driver.last_name[0]}."

            # 1. Дата
            ws.cell(row=row_idx, column=1, value=wb_obj.date.strftime('%d.%m.%Y'))  # A
            # 2. Наименование и место работы
            ws.cell(row=row_idx, column=2, value=name_place)                        # B
            # 3. Наличие ГСМ перед выездом
            cell = ws.cell(row=row_idx, column=3, value=rec.fuel_before_departure)  # C
            cell.fill = yellow_fill
            # 4. Время выезда
            ws.cell(row=row_idx, column=4, value=rec.departure_time.strftime('%H:%M'))  # D
            # 5. Время возвращения
            ws.cell(row=row_idx, column=5, value=rec.arrival_time.strftime('%H:%M'))    # E
            # 6. Показания спидометра перед выездом
            cell = ws.cell(row=row_idx, column=6, value=rec.odometer_before)       # F
            cell.fill = yellow_fill
            # 7. Пройдено км к месту работы и обратно
            ws.cell(row=row_idx, column=7, value=rec.distance_km)                  # G
            # 8. Израсходовано ГСМ, л (по пробегу)
            ws.cell(row=row_idx, column=8, value=rec.fuel_used_by_distance)        # H
            # 9–10. Наработка ПА, мин (с/без агрегата)
            ws.cell(row=row_idx, column=9, value=rec.time_with_pump)               # I
            ws.cell(row=row_idx, column=10, value=rec.time_without_pump)           # J
            # 11–12. Израсходовано ГСМ, л (с/без агрегата)
            ws.cell(row=row_idx, column=11, value=rec.fuel_used_with_pump)         # K
            ws.cell(row=row_idx, column=12, value=rec.fuel_used_without_pump)      # L
            # 13. Итого израсходовано ГСМ (факт)
            cell = ws.cell(row=row_idx, column=13, value=rec.fuel_used)            # M
            cell.fill = yellow_fill
            # 14. Израсходовано по норме
            cell = ws.cell(row=row_idx, column=14, value=rec.fuel_used_normal)     # N
            cell.fill = yellow_fill
            # 15. Получено ГСМ, л
            cell = ws.cell(row=row_idx, column=15, value=rec.fuel_refueled)        # O
            cell.fill = green_fill
            # 16. Наличие ГСМ при возвращении, л
            cell = ws.cell(row=row_idx, column=16, value=rec.fuel_on_return)       # P
            cell.fill = yellow_fill
            # 17. Показания спидометра при возвращении, км
            cell = ws.cell(row=row_idx, column=17, value=rec.odometer_after)       # Q
            cell.fill = yellow_fill

            # 18–19. Экономия/перерасход
            if rec.fuel_used_normal > rec.fuel_used:
                savings = rec.fuel_used_normal - rec.fuel_used
                overrun = Decimal('0.000')
            elif rec.fuel_used_normal < rec.fuel_used:
                overrun = rec.fuel_used - rec.fuel_used_normal
                savings = Decimal('0.000')
            else:
                savings = Decimal('0.000')
                overrun = Decimal('0.000')

            cell = ws.cell(row=row_idx, column=18, value=float(savings))           # R Экономия
            cell.fill = green_fill
            cell = ws.cell(row=row_idx, column=19, value=float(overrun))           # S Перерасход
            cell.fill = red_fill

            # Накапливаем итоги
            total_distance_km += rec.distance_km
            total_time_with_pump += rec.time_with_pump
            total_time_without_pump += rec.time_without_pump
            total_fuel_by_distance += rec.fuel_used_by_distance
            total_fuel_with_pump += rec.fuel_used_with_pump
            total_fuel_without_pump += rec.fuel_used_without_pump
            total_fuel_normal += rec.fuel_used_normal
            total_fuel_fact += rec.fuel_used
            total_fuel_refueled += rec.fuel_refueled
            total_savings += savings
            total_overrun += overrun

            row_idx += 1

        # ----- строка ИТОГО -----
        cell = ws.cell(row=row_idx, column=2, value="ИТОГО")        # Наименование/место можно оставить пустым
        cell.fill = yellow_fill

        ws.cell(row=row_idx, column=7, value=total_distance_km).fill = yellow_fill
        ws.cell(row=row_idx, column=9, value=total_time_with_pump).fill = yellow_fill
        ws.cell(row=row_idx, column=10, value=total_time_without_pump).fill = yellow_fill
        ws.cell(row=row_idx, column=8, value=float(total_fuel_by_distance)).fill = yellow_fill
        ws.cell(row=row_idx, column=11, value=float(total_fuel_with_pump)).fill = yellow_fill
        ws.cell(row=row_idx, column=12, value=float(total_fuel_without_pump)).fill = yellow_fill
        ws.cell(row=row_idx, column=13, value=float(total_fuel_fact)).fill = yellow_fill
        ws.cell(row=row_idx, column=14, value=float(total_fuel_normal)).fill = yellow_fill
        ws.cell(row=row_idx, column=15, value=float(total_fuel_refueled)).fill = green_fill
        ws.cell(row=row_idx, column=18, value=float(total_savings)).fill = green_fill
        ws.cell(row=row_idx, column=19, value=float(total_overrun)).fill = red_fill

        data_end_row = row_idx

        # ----- границы, шрифт, выравнивание для всех строк данных + итог -----
        for r in range(data_start_row, data_end_row + 1):
            for c in range(1, 21 + 1):  # колонки A..S
                cell = ws.cell(row=r, column=c)
                cell.border = thin_border
                cell.font = center_font
                cell.alignment = center_align

        # строку ИТОГО делаем жирной
        for c in range(1, 19 + 1):
            ws.cell(row=data_end_row, column=c).font = bold_font

        # ----- отдаём как файл -----
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = (f"Путевые листы пожарного автомобиля({car.number}) за период {from_date.strftime('%d.%m.%Y')}-{to_date.strftime('%d.%m.%Y')}.xlsx")
        quoted_filename = quote(filename)

        response = HttpResponse(
            output.read(),
            content_type=("application/vnd.openxmlformats-officedocument."
                          "spreadsheetml.sheet"),
        )
        response['Content-Disposition'] = f"attachment; filename*=UTF-8''{quoted_filename}"
        return response

class FireTruckWaybillRecordViewSet(SoftDeleteModelViewSet):
    queryset = FireTruckWaybillRecord.objects.select_related('fire_truck_waybill')
    serializer_class = FireTruckWaybillRecordSerializer